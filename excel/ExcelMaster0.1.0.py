#Project 3B ExcelMaster VER 0.0.1 (INCOMPLETE)
# Original Framework
# Custom spreadsheet interface program that allows for easy spreadsheet data storage integration. 
# Includes functions that can access or create new spreadsheets, those that can append, extract, or write new data, and more.
# If there is an internet connection, it can either create or append to an existing Google Sheets for cloud storage.

#Imports
from gspread.models import Spreadsheet
from openpyxl import load_workbook
import numpy as np
import random
import csv
import codecs
import urllib.request
import sys
import requests
from bs4 import BeautifulSoup
from pprint import pprint
import time
from sympy import *
from sys import exit
from datetime import datetime

#variable definition


#worksheet storage
workbooks = []

#addWorkbook: adds a new Workbook. Program will prevent the user from creating a spreadsheet with the same name to avoid confusion.
def addWorkbook():
    while True:
        try:
            name = (input("Enter the name of the new workbook:"))
        except (os.path.exists(name + '.xlsx')):
            print("ERROR: File with name already exists! Enter a different name:")
            continue
        else:
            new_Workbook = Workbook(name)
            workbooks[len(workbooks)] = name
            break
       
#addWorksheet: adds a worksheet to either an existing or new workbook.
def addWorksheet():
    workbookName = input("Enter the name of an existing or new workbook. Do NOT include .xlsx.")
    if(os.path.exists(workbookName + '.xlsx')):
        wb = load_workbook(filename = workbookName)
        worksheetName = input("Enter the name of the new worksheet to add to this workbook:")
        ws = wb.create_sheet(worksheet_Name)
    else:
        new_Workbook = Workbook(name)
        worksheetName = input("Enter the name of the new worksheet to add to this workbook:")
        ws = new_Workbook.create_sheet(worksheetName)
